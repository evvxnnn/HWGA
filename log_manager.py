"""
Log manager for creating and updating Excel log files from database entries
"""

import os
import pandas as pd
from datetime import datetime
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class LogManager:
    """Manages Excel log files for all modules"""
    
    def __init__(self):
        self.logs_dir = "logs"
        os.makedirs(self.logs_dir, exist_ok=True)
        self.db_path = "security_logs.db"
        self.init_log_files()
    
    def init_log_files(self):
        """Initialize all log files if they don't exist"""
        log_files = {
            "email_logs.xlsx": ["Date", "Time", "From", "Subject", "Category", "Site", "Priority", "Notes", "Logged By"],
            "phone_logs.xlsx": ["Date", "Time", "Caller", "Number", "Company", "Type", "Duration", "Summary", "Action Items", "Logged By"],
            "radio_logs.xlsx": ["Date", "Time", "Unit", "Location", "Type", "Message", "Priority", "Response", "Logged By"],
            "everbridge_logs.xlsx": ["Date", "Time", "Alert Type", "Severity", "Subject", "Message", "Recipients", "Response Rate", "Logged By"],
            "incident_logs.xlsx": ["Date", "Time", "Type", "Location", "Description", "Units Involved", "Status", "Resolution", "Logged By"],
            "facilities_logs.xlsx": ["Date", "Time", "Issue Type", "Location", "Description", "Priority", "Assigned To", "Status", "Logged By"],
            "data_request_logs.xlsx": ["Date", "Time", "Requester", "Department", "Request Type", "Description", "Due Date", "Status", "Logged By"],
            "badge_deactivation_logs.xlsx": ["Date", "Time", "Employee Name", "Badge Number", "Department", "Reason", "Effective Date", "Approved By", "Logged By"],
            "parking_logs.xlsx": ["Date", "Time", "Vehicle", "License Plate", "Location", "Issue", "Action Taken", "Officer", "Logged By"],
            "muster_logs.xlsx": ["Date", "Time", "Event Type", "Location", "Personnel Count", "Missing", "Accounted", "Notes", "Logged By"]
        }
        
        for filename, columns in log_files.items():
            filepath = os.path.join(self.logs_dir, filename)
            if not os.path.exists(filepath):
                df = pd.DataFrame(columns=columns)
                self.save_with_formatting(df, filepath)
    
    def save_with_formatting(self, df, filepath):
        """Save DataFrame to Excel with formatting"""
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Log Data')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Log Data']
            
            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border
    
    def add_email_log(self, email_data):
        """Add entry to email log"""
        filepath = os.path.join(self.logs_dir, "email_logs.xlsx")
        
        # Load existing data
        if os.path.exists(filepath):
            df = pd.read_excel(filepath)
        else:
            df = pd.DataFrame(columns=["Date", "Time", "From", "Subject", "Category", "Site", "Priority", "Notes", "Logged By"])
        
        # Add new row
        new_row = {
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Time": datetime.now().strftime("%H:%M:%S"),
            "From": email_data.get('from', 'Unknown'),
            "Subject": email_data.get('subject', ''),
            "Category": email_data.get('category', 'Other'),
            "Site": email_data.get('site', ''),
            "Priority": email_data.get('priority', 'Normal'),
            "Notes": email_data.get('notes', ''),
            "Logged By": os.environ.get('USERNAME', 'User')
        }
        
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_with_formatting(df, filepath)
        return df
    
    def add_phone_log(self, phone_data):
        """Add entry to phone log"""
        filepath = os.path.join(self.logs_dir, "phone_logs.xlsx")
        
        if os.path.exists(filepath):
            df = pd.read_excel(filepath)
        else:
            df = pd.DataFrame(columns=["Date", "Time", "Caller", "Number", "Company", "Type", "Duration", "Summary", "Action Items", "Logged By"])
        
        new_row = {
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Time": datetime.now().strftime("%H:%M:%S"),
            "Caller": phone_data.get('caller', 'Unknown'),
            "Number": phone_data.get('number', ''),
            "Company": phone_data.get('company', ''),
            "Type": phone_data.get('type', 'General'),
            "Duration": phone_data.get('duration', ''),
            "Summary": phone_data.get('summary', ''),
            "Action Items": phone_data.get('action_items', ''),
            "Logged By": os.environ.get('USERNAME', 'User')
        }
        
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_with_formatting(df, filepath)
        return df
    
    def add_radio_log(self, radio_data):
        """Add entry to radio log"""
        filepath = os.path.join(self.logs_dir, "radio_logs.xlsx")
        
        if os.path.exists(filepath):
            df = pd.read_excel(filepath)
        else:
            df = pd.DataFrame(columns=["Date", "Time", "Unit", "Location", "Type", "Message", "Priority", "Response", "Logged By"])
        
        new_row = {
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Time": datetime.now().strftime("%H:%M:%S"),
            "Unit": radio_data.get('unit', ''),
            "Location": radio_data.get('location', ''),
            "Type": radio_data.get('type', 'Dispatch'),
            "Message": radio_data.get('message', ''),
            "Priority": radio_data.get('priority', 'Normal'),
            "Response": radio_data.get('response', ''),
            "Logged By": os.environ.get('USERNAME', 'User')
        }
        
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_with_formatting(df, filepath)
        return df
    
    def add_everbridge_log(self, everbridge_data):
        """Add entry to Everbridge log"""
        filepath = os.path.join(self.logs_dir, "everbridge_logs.xlsx")
        
        if os.path.exists(filepath):
            df = pd.read_excel(filepath)
        else:
            df = pd.DataFrame(columns=["Date", "Time", "Alert Type", "Severity", "Subject", "Message", "Recipients", "Response Rate", "Logged By"])
        
        new_row = {
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Time": datetime.now().strftime("%H:%M:%S"),
            "Alert Type": everbridge_data.get('alert_type', ''),
            "Severity": everbridge_data.get('severity', 'Medium'),
            "Subject": everbridge_data.get('subject', ''),
            "Message": everbridge_data.get('message', ''),
            "Recipients": everbridge_data.get('recipients', ''),
            "Response Rate": everbridge_data.get('response_rate', ''),
            "Logged By": os.environ.get('USERNAME', 'User')
        }
        
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_with_formatting(df, filepath)
        return df
    
    def add_parking_log(self, parking_data):
        """Add entry to parking log"""
        filepath = os.path.join(self.logs_dir, "parking_logs.xlsx")
        
        if os.path.exists(filepath):
            df = pd.read_excel(filepath)
        else:
            df = pd.DataFrame(columns=["Date", "Time", "Vehicle", "License Plate", "Location", "Issue", "Action Taken", "Officer", "Logged By"])
        
        new_row = {
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Time": datetime.now().strftime("%H:%M:%S"),
            "Vehicle": parking_data.get('vehicle', ''),
            "License Plate": parking_data.get('license_plate', ''),
            "Location": parking_data.get('location', ''),
            "Issue": parking_data.get('issue', ''),
            "Action Taken": parking_data.get('action', ''),
            "Officer": parking_data.get('officer', ''),
            "Logged By": os.environ.get('USERNAME', 'User')
        }
        
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self.save_with_formatting(df, filepath)
        return df
    
    def get_recent_logs(self, log_type, limit=10):
        """Get recent log entries for display"""
        log_files = {
            "email": "email_logs.xlsx",
            "phone": "phone_logs.xlsx",
            "radio": "radio_logs.xlsx",
            "everbridge": "everbridge_logs.xlsx",
            "parking": "parking_logs.xlsx",
            "incident": "incident_logs.xlsx",
            "facilities": "facilities_logs.xlsx",
            "badge": "badge_deactivation_logs.xlsx",
            "muster": "muster_logs.xlsx",
            "data_request": "data_request_logs.xlsx"
        }
        
        filename = log_files.get(log_type)
        if not filename:
            return pd.DataFrame()
        
        filepath = os.path.join(self.logs_dir, filename)
        if os.path.exists(filepath):
            df = pd.read_excel(filepath)
            # Return most recent entries
            return df.tail(limit)
        return pd.DataFrame()
    
    def sync_from_database(self):
        """Sync all logs from database to Excel files"""
        if not os.path.exists(self.db_path):
            return
        
        conn = sqlite3.connect(self.db_path)
        
        # Sync email logs
        try:
            email_df = pd.read_sql_query("""
                SELECT 
                    date(timestamp) as Date,
                    time(timestamp) as Time,
                    sender as 'From',
                    subject as Subject,
                    category as Category,
                    site_code as Site,
                    priority as Priority,
                    notes as Notes,
                    logged_by as 'Logged By'
                FROM email_logs
                ORDER BY timestamp DESC
            """, conn)
            if not email_df.empty:
                self.save_with_formatting(email_df, os.path.join(self.logs_dir, "email_logs.xlsx"))
        except:
            pass
        
        # Sync phone logs
        try:
            phone_df = pd.read_sql_query("""
                SELECT 
                    date(timestamp) as Date,
                    time(timestamp) as Time,
                    caller_name as Caller,
                    phone_number as Number,
                    company as Company,
                    call_type as Type,
                    duration as Duration,
                    summary as Summary,
                    action_items as 'Action Items',
                    logged_by as 'Logged By'
                FROM phone_logs
                ORDER BY timestamp DESC
            """, conn)
            if not phone_df.empty:
                self.save_with_formatting(phone_df, os.path.join(self.logs_dir, "phone_logs.xlsx"))
        except:
            pass
        
        # Sync radio logs
        try:
            radio_df = pd.read_sql_query("""
                SELECT 
                    date(timestamp) as Date,
                    time(timestamp) as Time,
                    unit_number as Unit,
                    location as Location,
                    dispatch_type as Type,
                    message as Message,
                    priority as Priority,
                    response_action as Response,
                    logged_by as 'Logged By'
                FROM radio_logs
                ORDER BY timestamp DESC
            """, conn)
            if not radio_df.empty:
                self.save_with_formatting(radio_df, os.path.join(self.logs_dir, "radio_logs.xlsx"))
        except:
            pass
        
        conn.close()


# Global log manager instance
log_manager = LogManager()